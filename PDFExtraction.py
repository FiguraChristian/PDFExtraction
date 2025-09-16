# Der folgende Code holt sich aus allen PDF Dateien in einem definierten Ordner den jeweiligen Gesamtbetrag
# und speichert diese in einem Excel-File "Rechnungsbetraege.xlsx".
# Um die Wiederverwendung zu ermöglichen, überschriebt das Skript bei jeder Ausführung die bestehenden Daten.


# Import der benötigten Bibliotheken
import os
import re
from pypdf import PdfReader
import openpyxl

# Zielverzeichnis, welches durchsucht werden soll
# Sollte in der Praxis generisch sein
SOURCE = "C:\\Users\\chris\\Documents\\PDFsZumAuslesen"

# Blanko-Listen, um Beträge zu extrahieren und um iterieren zu können
list_of_elements = []
list_of_amounts = []
final_list = []

# Iteration im Zielverzeichnis
if os.path.exists(SOURCE):
    for element in os.listdir(SOURCE):
        if element.endswith(".pdf"):
            # Dateipfad ermitteln
            element_path = os.path.join(SOURCE, element)
            # Datei einlesen
            pdf = PdfReader(element_path)
            # Rechnungsbetrag der ersten Seite "rausholen"
            pdf_page = pdf.pages[0]
            # Textextraktion
            content_pdf = pdf_page.extract_text()

            # Muster für die qualifizierte Textsuche
            # muss für Praxis an Umgebung angepasst werden
            pattern = re.compile(r"[0-9]+,[0-9]+")
            amounts_as_string = pattern.findall(content_pdf)

            # Konvertiert die gefundenen Strings in Floats, da "findall" immer einen String ausgibt und nicht das Objekt
            amounts_as_float = []
            for amount in amounts_as_string:
                # Entfernt Tausendertrennzeichen und ersetzt Komma durch Punkt für float-Konvertierung
                cleaned_amount = amount.replace(".", "").replace(",", ".")
                amounts_as_float.append(float(cleaned_amount))

            # Findet den größten Betrag in der aktuellen PDF und fügt ihn zur Liste hinzu
            if amounts_as_float:
                max_amount = max(amounts_as_float)
                final_list.append((element, max_amount))
                print(f"Der Gesamtbetrag in '{element}' ist: {max_amount}")
            else:
                print(f"Keine Beträge in '{element}' gefunden.")


    print("\nDie Liste der größten Beträge pro PDF:")
    print(final_list)

# Reporting

    if final_list:
        # Erstelle eine neue Arbeitsmappe
        wb = openpyxl.Workbook()
        # Wähle das aktive Arbeitsblatt aus
        ws = wb.active
        ws.title = "Rechnungsbetraege"

        # Schreibe die Spaltenüberschriften
        ws['A1'] = "Dateiname"
        ws['B1'] = "Rechnungsbetrag in €"

        # Iteriere durch die final_list und schreibe die Daten in das Arbeitsblatt
        for row_num, (filename, amount) in enumerate(final_list, start=2):
            ws[f'A{row_num}'] = filename
            ws[f'B{row_num}'] = amount

        # Definiere den Speicherpfad und Dateinamen für die Excel-Datei
        excel_path = os.path.join(SOURCE, "Rechnungsbetraege.xlsx")

        try:
            # Speichere die Arbeitsmappe
            wb.save(excel_path)
            print(f"\nExcel-Datei erfolgreich gespeichert unter: {excel_path}")
        except Exception as e:
            print(f"\nFehler beim Speichern der Excel-Datei: {e}")
    else:
        print("\nKeine Rechnungsbeträge gefunden, es wird keine Excel-Datei erstellt.")